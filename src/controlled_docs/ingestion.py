from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree
from zipfile import ZipFile

from sqlalchemy import select
from sqlalchemy.orm import Session

from controlled_docs.audit import log_event
from controlled_docs.chunking import chunk_text
from controlled_docs.config import AppConfig
from controlled_docs.embeddings import Embedder
from controlled_docs.models import Chunk, Document, DocumentStatus, DocumentVersion
from controlled_docs.normalization import NormalizedMetadata, normalize_metadata
from controlled_docs.utils import parse_date, read_text_file, sha256_text, stable_uuid

LOGGER = logging.getLogger(__name__)


@dataclass(slots=True)
class DocumentInput:
    document_key: str
    title: str
    doc_type: str
    jurisdiction: str
    owner: str
    status: str
    version_number: int
    effective_date: str
    file_path: Path

    @property
    def normalized(self) -> NormalizedMetadata:
        return normalize_metadata(
            document_key=self.document_key,
            title=self.title,
            doc_type=self.doc_type,
            jurisdiction=self.jurisdiction,
            owner=self.owner,
            status=self.status,
        )


@dataclass(slots=True)
class IngestionSummary:
    documents_created: int = 0
    versions_created: int = 0
    chunks_created: int = 0
    skipped_versions: int = 0


def load_register_entries(path: Path) -> list[DocumentInput]:
    register_root = path.resolve().parent.parent
    headers, rows = _read_xlsx_rows(path)
    entries: list[DocumentInput] = []
    for row in rows:
        if not any(row.values()):
            continue
        payload = row
        file_path = Path(str(payload["file_path"]))
        resolved_path = file_path if file_path.is_absolute() else (register_root / file_path).resolve()
        entries.append(
            DocumentInput(
                document_key=str(payload["document_key"]),
                title=str(payload["title"]),
                doc_type=str(payload["doc_type"]),
                jurisdiction=str(payload["jurisdiction"]),
                owner=str(payload["owner"]),
                status=str(payload["status"]),
                version_number=int(payload["version_number"]),
                effective_date=str(payload["effective_date"]),
                file_path=resolved_path,
            )
        )
    return entries


def _read_xlsx_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path)
        sheet = workbook["document_register"]
        headers = [str(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        rows: list[dict[str, str]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            rows.append({header: str(value) for header, value in zip(headers, row, strict=True)})
        return headers, rows
    except ImportError:
        return _read_xlsx_rows_stdlib(path)


def _read_xlsx_rows_stdlib(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read("xl/worksheets/sheet1.xml"))
    row_nodes = root.findall(".//main:row", namespace)
    parsed_rows: list[list[str]] = []
    for row_node in row_nodes:
        values: list[str] = []
        for cell in row_node.findall("main:c", namespace):
            inline = cell.find("main:is/main:t", namespace)
            value_node = cell.find("main:v", namespace)
            if inline is not None:
                values.append(inline.text or "")
            elif value_node is not None:
                values.append(value_node.text or "")
            else:
                values.append("")
        parsed_rows.append(values)
    headers = parsed_rows[0]
    rows = [
        {header: row[index] for index, header in enumerate(headers)}
        for row in parsed_rows[1:]
        if row
    ]
    return headers, rows


def parse_frontmatter(text: str) -> dict[str, str]:
    metadata: dict[str, str] = {}
    for line in text.splitlines():
        if not line.strip():
            break
        if ":" not in line:
            continue
        key, value = line.split(":", 1)
        metadata[key.strip().lower().replace(" ", "_")] = value.strip()
    return metadata


def load_folder_entries(path: Path) -> list[DocumentInput]:
    entries: list[DocumentInput] = []
    for file_path in sorted(path.rglob("*")):
        if file_path.suffix.lower() not in {".txt", ".md"}:
            continue
        metadata = parse_frontmatter(read_text_file(file_path))
        entries.append(
            DocumentInput(
                document_key=metadata["document_key"],
                title=metadata["title"],
                doc_type=metadata["doc_type"],
                jurisdiction=metadata["jurisdiction"],
                owner=metadata["owner"],
                status=metadata["status"],
                version_number=int(metadata["version_number"]),
                effective_date=metadata["effective_date"],
                file_path=file_path,
            )
        )
    return entries


class IngestionService:
    def __init__(self, session: Session, config: AppConfig, embedder: Embedder):
        self.session = session
        self.config = config
        self.embedder = embedder

    def ingest_register(self, path: Path, *, actor: str = "system") -> IngestionSummary:
        return self.ingest_entries(load_register_entries(path), actor=actor)

    def ingest_folder(self, path: Path, *, actor: str = "system") -> IngestionSummary:
        return self.ingest_entries(load_folder_entries(path), actor=actor)

    def ingest_entries(self, entries: Iterable[DocumentInput], *, actor: str = "system") -> IngestionSummary:
        summary = IngestionSummary()
        for entry in entries:
            summary = self._ingest_single(entry, actor=actor, summary=summary)
        LOGGER.info("Ingestion summary: %s", summary)
        return summary

    def _ingest_single(
        self, entry: DocumentInput, *, actor: str, summary: IngestionSummary
    ) -> IngestionSummary:
        metadata = entry.normalized
        text = read_text_file(entry.file_path)
        checksum = sha256_text(text)

        document_id = stable_uuid(f"{metadata.document_key}|document")
        version_id = stable_uuid(f"{metadata.document_key}|version|{entry.version_number}")

        document = self.session.get(Document, document_id)
        if document is None:
            document = Document(
                document_id=document_id,
                title=metadata.title,
                doc_type=metadata.doc_type,
                jurisdiction=metadata.jurisdiction,
                owner=metadata.owner,
                status=metadata.status,
                current_version_id=None,
            )
            self.session.add(document)
            summary.documents_created += 1
            log_event(
                self.session,
                actor=actor,
                action="DOCUMENT_CREATED",
                object_type="document",
                object_id=str(document_id),
                details={
                    "title": metadata.title,
                    "doc_type": metadata.doc_type,
                    "jurisdiction": metadata.jurisdiction,
                },
            )
        elif document.status != metadata.status:
            old_status = document.status.value
            document.status = metadata.status
            log_event(
                self.session,
                actor=actor,
                action="STATUS_TRANSITION",
                object_type="document",
                object_id=str(document_id),
                details={"from": old_status, "to": metadata.status.value},
            )

        existing_version = self.session.get(DocumentVersion, version_id)
        if existing_version is not None:
            summary.skipped_versions += 1
            log_event(
                self.session,
                actor=actor,
                action="VERSION_SKIPPED",
                object_type="document_version",
                object_id=str(version_id),
                details={"reason": "already_exists", "source_path": str(entry.file_path)},
            )
            return summary

        previous_version = self.session.scalar(
            select(DocumentVersion)
            .where(DocumentVersion.document_id == document_id)
            .order_by(DocumentVersion.version_number.desc())
            .limit(1)
        )
        version = DocumentVersion(
            version_id=version_id,
            document_id=document_id,
            version_number=entry.version_number,
            effective_date=parse_date(entry.effective_date),
            supersedes_version_id=previous_version.version_id if previous_version else None,
            checksum_sha256=checksum,
            source_path=str(entry.file_path),
        )
        self.session.add(version)
        summary.versions_created += 1

        chunks = chunk_text(
            text,
            chunk_size=self.config.chunk_size,
            chunk_overlap=self.config.chunk_overlap,
        )
        embeddings = self.embedder.embed_texts(chunks) if chunks else []
        for chunk_index, (chunk_text_value, embedding) in enumerate(zip(chunks, embeddings, strict=True)):
            chunk_id = stable_uuid(f"{version_id}|chunk|{chunk_index}")
            self.session.add(
                Chunk(
                    chunk_id=chunk_id,
                    version_id=version_id,
                    chunk_index=chunk_index,
                    text=chunk_text_value,
                    embedding=embedding,
                    metadata_json={
                        "title": metadata.title,
                        "jurisdiction": metadata.jurisdiction,
                        "doc_type": metadata.doc_type,
                        "status": metadata.status.value,
                        "version_number": entry.version_number,
                        "effective_date": str(parse_date(entry.effective_date)),
                    },
                )
            )
            summary.chunks_created += 1

        if previous_version is None or entry.version_number >= previous_version.version_number:
            document.current_version_id = version_id
            document.status = metadata.status

        log_event(
            self.session,
            actor=actor,
            action="INGEST_VERSION",
            object_type="document_version",
            object_id=str(version_id),
            details={
                "document_id": str(document_id),
                "version_number": entry.version_number,
                "checksum_sha256": checksum,
                "chunk_count": len(chunks),
                "source_path": str(entry.file_path),
            },
        )
        return summary
